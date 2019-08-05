# -*- coding: utf-8 -*-
'''
    Arquivo com funcoes para trabalhar com planilhas XLSX


Criador : MuriloCunha
Data    : 04/08/2019

LOG de alteracoes:
-------------------------------------------------
04/08/2019- Criacao deste arquivo

-------------------------------------------------
'''


from datetime import date, datetime
import os
from openpyxl import Workbook
import openpyxl
from funArquivoTexto import ArquivoXlsx


def as_text(value):
    if value is None:
        return ''
    else:
        return str(value)


def GravarResultadoEmPlanilha(lstDadosAGravar, lstCabecalho, strVendedor):
    '''
        Gravacao dos dados no XLS
        Acrescentar os dados recuperados do banco na planilha
    '''
    # Criando a pasta de trabalho
    book = Workbook()

    # Ativando a planilha da pasta de trabalho
    sheet = book.active

    # Acrescentar, na primeira linha, o cabecalho de colunas
    sheet.append(lstCabecalho)
    
    c = 1
    for linha in lstDadosAGravar:
        print('[INFO] Vendedor: {}\t-Inserindo linha: {}'.format(
            strVendedor, c))
        #
        # A linha abaixo insere EXATAMENTE como o dado se apresenta no PRINT
        # Se for um NUMBER e o separador de milhar for "." (ponto), a planilha
        # nao fará os calculos como numericos, apenas como STRING.
        # sheet.cell(row=idx+2, column=c+2).value = str(celula)
        
        #
        # Na linha abaixo, a biblioteca se encarrega de formatar o tipo correto
        # para cada coluna (STRING, NUMBER, DATE...) de acordo com o dado.
        # Entao, esta forma é preferida para inserir varias info sem se preocupar
        # com os tipos
        sheet.append(linha)
        c += 1

        # Formatar coluna especiais com o formato DATA e CONTABIL
        strFormatoDataAbreviada = 'DD/MM/YYYY'
        strFormatoContabil = 'R$ #,###,##0.00;[RED]-R$ #,###,##0.00;R$ 0.00'
        sheet['A'+str(c)].number_format = strFormatoDataAbreviada
        sheet['K'+str(c)].number_format = strFormatoContabil
        sheet['L'+str(c)].number_format = strFormatoContabil
        sheet['M'+str(c)].number_format = strFormatoContabil

    # Adicionar o TOTAL na coluna L
    c += 2
    sheet['K'+str(c)] = 'Total'
    sheet['L'+str(c)].number_format = strFormatoContabil
    sheet['L'+str(c)] = '=SUM(L1:L'+str(c-1)+')'
    
    # Ajustar largura das colunas para mostrar todo o conteudo
    for column_cells in sheet.columns:
        length = max(len(str(cell.value) or '') for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = length    
    
    # Colocando filtro nas colunas
    #sheet.auto_filter.ref = sheet.dimensions
    
    # Uma segunda forma de colocar filtro nas colunas
    maxcolumnletter = openpyxl.utils.get_column_letter(sheet.max_column)
    sheet.auto_filter.ref = 'A1:'+maxcolumnletter+str(len(sheet['A']))

    sheet.auto_filter.add_filter_column(0, ['30/07/2019'])
    sheet.auto_filter.add_sort_condition("B2:B500")


    
    #print(agora.year, '-', agora.month, '-', agora.day, ' | ',
    #      agora.hour, ':', agora.minute, ':', agora.second)
    datAgora = datetime.now()
    strData = datAgora.strftime('%Y%m%d')
    strHora = datAgora.strftime('%H%M%S')
    strArqXlsx = 'Cubo_8003_{}_{}_{}.xlsx'.format(
        strVendedor, strData, strHora)

    # Criar classe para manipulacao de arquivos XLSX
    arqXlsx = ArquivoXlsx('./Resultado', strArqXlsx)
    
    # Criar diretorio se nao existir
    arqXlsx.PrepararDiretorioParaGravacao()

    # Apagar arquivo se este existir
    arqXlsx.ApagarArquivoExistente()
    
    strCaminhoCompleto = os.path.join('./Resultado', strArqXlsx)
    
    # Gravando o arquivo XLS
    book.save(strCaminhoCompleto)

    # FIM - Gravacao do XLS
    #

    print('Arquivo(s) criado(s) com sucesso:', strArqXlsx, '\n')

